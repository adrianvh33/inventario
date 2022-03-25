from django.shortcuts import render
from django.http import Http404,HttpResponseRedirect,HttpResponse
from django.urls import reverse
from .models import Producto
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView,CreateView, UpdateView
from django.views.generic.edit import DeleteView
from .forms import ProductosForms, ImportarForm, ContadorForm
import datetime
from openpyxl import load_workbook,Workbook
from django.db import transaction
from django.db.models import Q

# Create your views here.



class ProductosDeleteView(LoginRequiredMixin,DeleteView):
    model = Producto
    success_url = '/productos'
    template_name = 'productos/productos_delete.html'
    login_url = '/'


class ProductosUpdateView(LoginRequiredMixin,UpdateView):
    model= Producto
    template_name = 'productos/productos_form.html'
    success_url = '/productos'
    form_class = ProductosForms
    login_url = '/'
    
    def form_valid(self, form):
        form.instance.editado = datetime.datetime.now()
        return super().form_valid(form)


class ProductosCreateView(LoginRequiredMixin,CreateView):
    model= Producto
    template_name = 'productos/productos_form.html'
    success_url = '/productos'
    form_class = ProductosForms
    login_url = '/'
    
    def form_valid(self, form):
        form.instance.editado = datetime.datetime.now()
        return super().form_valid(form)

class ProductosListView(LoginRequiredMixin,ListView):
    model = Producto
    context_object_name = 'productos'
    template_name = 'productos/productos_list.html'
    login_url = '/'
    productos = []

    def get_queryset(self, *args, **kwargs):
        qs = super().get_queryset(*args, **kwargs)
        query = self.request.GET.get('q')
        if query:
            return qs.filter(Q(referencia=query)|Q(nombre__contains=query))
        return qs


class ProductoDetailView(LoginRequiredMixin,UpdateView):
    model = Producto
    template_name = 'productos/producto_detalle.html'
    success_url = '/productos'
    form_class = ContadorForm
    login_url = '/'

    def form_valid(self, form):
        form.save()
        return super(ProductoDetailView, self).form_valid(form)


def importar(request):
    editado = datetime.datetime.now()
    sexos = ('hombre','dama')
    if request.method == 'POST':
        filled_form = ImportarForm(request.POST, request.FILES)
        if filled_form.is_valid():
            file = request.FILES['file']
            recontar = filled_form.cleaned_data['recontar']
            print(recontar)
            wb = load_workbook(file)
            sheet = wb.active
            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, min_col=1):
                    referencia = row[0].value
                    nombre = row[1].value 
                    cantidadSistema = row[2].value
                    if row[3].value:
                        precio = row[3].value
                    else:
                        precio = 0
                    
                    if Producto.objects.filter(referencia=referencia).exists():
                        prod = Producto.objects.get(referencia=referencia)
                        if recontar == 'noContar':                            
                            prod.cantidadSistema=cantidadSistema
                            prod.enTienda = 0
                            prod.enBloque2 = 0
                            prod.enBloque5 = 0
                            prod.save()
                        else:
                            prod.cantidadSistema=cantidadSistema
                            prod.save()
                    else:
                        Producto.objects.create(referencia=referencia,nombre=nombre,cantidadSistema=cantidadSistema,precio=precio,editado=editado)
            return HttpResponseRedirect(reverse('productos.list'))
        else:
            form = ImportarForm()
            print(filled_form.errors)
    else:
        form = ImportarForm()
    return render(request,'productos/importar.html', {'form':form})


def reporte(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte.xlsx'
    wb = Workbook() 
    products_list = Producto.objects.all()
    sheet = wb.active
    sheet.cell(row=1,column=1,value='Referencia')
    sheet.cell(row=1,column=2,value="nombre")
    sheet.cell(row=1,column=3,value="Cantidad_Sistema")
    sheet.cell(row=1,column=4,value="Total")
    sheet.cell(row=1,column=5,value="En_tienda")
    sheet.cell(row=1,column=6,value="En_bloque_2")
    sheet.cell(row=1,column=7,value="En_bloque_5")
    i = 2
    for p in products_list:
        sheet.cell(row=i,column=1,value=p.referencia)
        sheet.cell(row=i,column=2,value=p.nombre)
        sheet.cell(row=i,column=3,value=p.cantidadSistema)
        sheet.cell(row=i,column=4,value=p.enTienda + p.enBloque2 +p.enBloque5)
        sheet.cell(row=i,column=5,value=p.enTienda)
        sheet.cell(row=i,column=6,value=p.enBloque2)
        sheet.cell(row=i,column=7,value=p.enBloque5)
        i += 1
    wb.save(response)     
    return response

def deleteAll(request):
    if request.method == 'POST':
        Producto.objects.all().delete()
        return HttpResponseRedirect(reverse('productos.list'))
    return render(request,'productos/productos_delete_all.html', {})